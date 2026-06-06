"""Build / update the master Excel workbook for narrative-sample coding.

Layout
------
Row 1: general coding rule (merged across all key columns).
Row 2: section name for each key.
Row 3: JSON key (column id).
Row 4: full Arabic description (the JSON value).
Row 5: coding scale for that key (binary / qualitative / numeric).
Row 6+: one row per recorded session.

Metadata columns (session id, child name, date, audio dir, notes) are inserted
*before* the data columns so analysts can identify each row at a glance.

The scoring scale per column is inferred from the field name:

* keys ending with ``_correct`` / ``_accuracy``        → binary 0/1
* keys ending with ``_score`` (excluding totals)       → qualitative 0/1/2
* keys ending with ``_count`` / ``_tokens`` / ``_raw`` → integer count
* keys ending with ``_percentage`` / ``_ratio`` / ``_density`` / ``TTR_*`` → float
* keys that hold a list (``_list``) or freeform text   → text
* totals (``*_total_score`` / ``macro_*``)             → numeric sum

The general coding rule (provided by the user) is written verbatim in row 1.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .questions import load_all_keys


GENERAL_RULE = (
    "قاعدة الترميز العامة: تُرمّز البنود الثنائية كالتالي: "
    "0 = غير موجود أو غير صحيح، 1 = موجود أو صحيح. "
    "وتُرمّز البنود النوعية أو المركبة كالتالي: "
    "0 = غائب، 1 = جزئي أو محدود، 2 = واضح أو مكتمل. "
    "أي اختلاف في سلم الترميز يجب أن يُذكر صراحة داخل البند."
)

METADATA_COLUMNS: List[str] = [
    "session_id",
    "child_name",
    "child_age",
    "child_gender",
    "child_group",
    "date_iso",
    "audio_dir",
    "notes",
]


def infer_scale(key: str) -> str:
    """Return a short human-readable scale label for a JSON key."""
    k = key.lower()
    if k.endswith("_total_score") or k.startswith("macro_") or "total" in k:
        return "مجموع رقمي"
    if k.endswith("_score"):
        return "نوعي: 0 = غائب، 1 = جزئي، 2 = واضح/مكتمل"
    if k.endswith("_correct") or k.endswith("_accuracy"):
        return "ثنائي: 0 = غير صحيح، 1 = صحيح"
    if k.endswith("_list") or k.endswith("_examples_optional") or k == "justification_type":
        return "نص حر"
    if (
        k.endswith("_count")
        or k.endswith("_tokens")
        or k.endswith("_raw")
        or k.endswith("_tnw")
        or k.endswith("_ndw")
        or k.endswith("_count_optional")
    ):
        return "عدد صحيح"
    if (
        k.endswith("_percentage")
        or k.endswith("_ratio")
        or k.endswith("_density")
        or k.startswith("ttr_")
        or k.endswith("_mlu_words_optional")
        or "mlu" in k
    ):
        return "نسبة / كسر عشري"
    if k.endswith("_response") or k.endswith("_speech_correct"):
        return "ثنائي: 0 = غير صحيح، 1 = صحيح"
    return "حسب الوصف"


def default_value_for(key: str) -> Optional[object]:
    """Default cell value when a fresh session row is appended.

    Returns ``None`` so the analyst can fill the cell. We intentionally do not
    auto-score answers – coding is a manual / hybrid task per the protocol.
    """
    return None


# ----------------------------------------------------------------------
# Header construction
# ----------------------------------------------------------------------
_HEADER_FILL = PatternFill("solid", fgColor="FFEFD5")
_RULE_FILL = PatternFill("solid", fgColor="D9E1F2")
_SECTION_FILL = PatternFill("solid", fgColor="E2EFDA")
_SCALE_FILL = PatternFill("solid", fgColor="FFF2CC")
_META_FILL = PatternFill("solid", fgColor="F8CBAD")


def _write_headers(ws, keys_meta: List[Dict[str, str]]) -> None:
    """Populate the 5 header rows of the worksheet."""
    total_cols = len(METADATA_COLUMNS) + len(keys_meta)

    # Row 1 – general coding rule (merged over key columns only).
    ws.cell(row=1, column=1, value="القاعدة العامة:").font = Font(bold=True)
    ws.cell(row=1, column=1).fill = _RULE_FILL
    if total_cols > len(METADATA_COLUMNS):
        ws.merge_cells(
            start_row=1,
            start_column=len(METADATA_COLUMNS) + 1,
            end_row=1,
            end_column=total_cols,
        )
        rule_cell = ws.cell(
            row=1, column=len(METADATA_COLUMNS) + 1, value=GENERAL_RULE
        )
        rule_cell.alignment = Alignment(wrap_text=True, vertical="center")
        rule_cell.fill = _RULE_FILL
        rule_cell.font = Font(bold=True)

    # Metadata column headers (rows 2-3 merged, row 4-5 left blank).
    for idx, name in enumerate(METADATA_COLUMNS, start=1):
        cell = ws.cell(row=2, column=idx, value=name)
        cell.font = Font(bold=True)
        cell.fill = _META_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=2, start_column=idx, end_row=3, end_column=idx)
        ws.cell(row=4, column=idx).fill = _META_FILL
        ws.cell(row=5, column=idx, value="—").alignment = Alignment(horizontal="center")
        ws.cell(row=5, column=idx).fill = _META_FILL

    # Key columns: section / key / description / scale.
    for offset, meta in enumerate(keys_meta):
        col = len(METADATA_COLUMNS) + 1 + offset

        section_cell = ws.cell(row=2, column=col, value=meta["section"])
        section_cell.fill = _SECTION_FILL
        section_cell.font = Font(bold=True)
        section_cell.alignment = Alignment(wrap_text=True, horizontal="center")

        topic_cell = ws.cell(row=3, column=col, value=meta["topic"])
        topic_cell.fill = _SECTION_FILL
        topic_cell.alignment = Alignment(wrap_text=True, horizontal="center")

        key_cell = ws.cell(row=4, column=col, value=meta["key"])
        key_cell.fill = _HEADER_FILL
        key_cell.font = Font(bold=True)
        key_cell.alignment = Alignment(horizontal="center")

        desc_cell = ws.cell(row=5, column=col, value=meta["description"])
        desc_cell.alignment = Alignment(wrap_text=True, vertical="top")

        scale_cell = ws.cell(row=6, column=col, value=infer_scale(meta["key"]))
        scale_cell.fill = _SCALE_FILL
        scale_cell.alignment = Alignment(wrap_text=True, horizontal="center")

    # Heading row for data starts at row 7 (after the scale row).
    ws.freeze_panes = ws.cell(row=7, column=len(METADATA_COLUMNS) + 1)

    # Column widths.
    for idx in range(1, total_cols + 1):
        letter = get_column_letter(idx)
        if idx <= len(METADATA_COLUMNS):
            ws.column_dimensions[letter].width = 20
        else:
            ws.column_dimensions[letter].width = 28
    # Row heights for the description row.
    ws.row_dimensions[1].height = 60
    ws.row_dimensions[5].height = 120
    ws.row_dimensions[6].height = 40


def _data_start_row() -> int:
    # Row 1 rule, row 2 section, row 3 topic, row 4 key,
    # row 5 description, row 6 scale, row 7+ data.
    return 7


# ----------------------------------------------------------------------
# Public entrypoints
# ----------------------------------------------------------------------
def create_workbook(json_path: Path, output_path: Path) -> Path:
    """Create a fresh workbook with all headers ready for data entry."""
    keys_meta = load_all_keys(json_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "narrative_coding"
    ws.sheet_view.rightToLeft = True
    _write_headers(ws, keys_meta)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def append_session(
    json_path: Path,
    workbook_path: Path,
    session_meta: Dict[str, str],
    answers: Optional[Dict[str, object]] = None,
) -> int:
    """Append a session row to the workbook (creating it if missing).

    ``answers`` may contain any subset of the JSON keys; missing keys are left
    blank so the analyst can fill them in later. Returns the 1-based row number
    that was appended.
    """
    if not workbook_path.exists():
        create_workbook(json_path, workbook_path)

    keys_meta = load_all_keys(json_path)
    wb = load_workbook(workbook_path)
    ws = wb.active

    # Find first empty data row.
    row = _data_start_row()
    while ws.cell(row=row, column=1).value not in (None, ""):
        row += 1

    # Write metadata columns.
    for idx, name in enumerate(METADATA_COLUMNS, start=1):
        ws.cell(row=row, column=idx, value=session_meta.get(name, ""))

    # Write key columns.
    answers = answers or {}
    for offset, meta in enumerate(keys_meta):
        col = len(METADATA_COLUMNS) + 1 + offset
        if meta["key"] in answers:
            ws.cell(row=row, column=col, value=answers[meta["key"]])
        else:
            ws.cell(row=row, column=col, value=default_value_for(meta["key"]))

    wb.save(workbook_path)
    return row


def make_session_meta(
    session_id: str,
    child_name: str,
    child_age: str,
    audio_dir: Path,
    notes: str = "",
    child_gender: str = "",
    child_group: str = "",
) -> Dict[str, str]:
    return {
        "session_id": session_id,
        "child_name": child_name,
        "child_age": child_age,
        "child_gender": child_gender,
        "child_group": child_group,
        "date_iso": datetime.now().isoformat(timespec="seconds"),
        "audio_dir": str(audio_dir),
        "notes": notes,
    }


if __name__ == "__main__":
    here = Path(__file__).resolve().parent.parent
    out = create_workbook(
        here / "data" / "extracted_keys.json",
        here / "output" / "coding_template.xlsx",
    )
    print(f"Wrote {out}")
